#!/usr/bin/env python3
"""
CSI TUM Report Generator
Generates Excel reports with transparent calculations using Excel formulas.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Default data path
DEFAULT_DATA_PATH = r"\\p1fs002\CSI - Technical Services\TUM Data for Accountants FY27"


def load_site_data(csv_path: str) -> pd.DataFrame:
    """Load and parse site CSV data."""
    df = pd.read_csv(csv_path)
    df = df.dropna(how='all')
    return df


def col_letter(col_num):
    """Convert column number to Excel letter (1=A, 2=B, etc.)"""
    return get_column_letter(col_num)


def create_csi_tum_report(site_name: str,
                          forecast_csv_path: str,
                          output_path: str,
                          historical_csv_path: str = None,
                          contract_tonnes: float = None,
                          target_tph: float = None,
                          lump_split: float = 0.50,
                          fines_split: float = 0.50,
                          commodity: str = "iron-ore"):
    """
    Generate a CSI TUM Report in Excel format with transparent calculations.
    All totals and averages use Excel formulas for full auditability.
    """

    # Load forecast data
    df_forecast = load_site_data(forecast_csv_path)
    num_forecast_months = len(df_forecast)

    # Load historical data if provided
    df_historical = None
    num_hist_months = 0
    if historical_csv_path:
        try:
            df_historical = load_site_data(historical_csv_path)
            num_hist_months = len(df_historical)
        except Exception as e:
            print(f"Warning: Could not load historical data: {e}")

    # Create workbook
    wb = Workbook()

    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=11, bold=True)
    title_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    section_font = Font(name='Arial', size=12, bold=True, color='1F4E79')
    value_font = Font(name='Arial', size=10)
    formula_font = Font(name='Arial', size=10, color='0066CC')
    kpi_value_font = Font(name='Arial', size=16, bold=True, color='1F4E79')

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    light_fill = PatternFill(start_color='D6E3F8', end_color='D6E3F8', fill_type='solid')
    forecast_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    historical_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    calc_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # =========================================================================
    # SHEET 1: FORECAST DATA (Raw data with formulas)
    # =========================================================================
    ws_forecast = wb.active
    ws_forecast.title = "Forecast Data"

    # Header
    ws_forecast.merge_cells('A1:P1')
    ws_forecast['A1'] = f"FY27 FORECAST DATA - {site_name}"
    ws_forecast['A1'].font = header_font
    ws_forecast['A1'].fill = header_fill
    ws_forecast['A1'].alignment = Alignment(horizontal='center')

    ws_forecast['A2'] = f"Generated: {datetime.now().strftime('%d/%m/%Y')}"
    ws_forecast['A2'].font = value_font

    # Column headers
    forecast_headers = ['Month', 'Tonnes', 'Run Hours', 'TPH', 'Availability', 'Utilisation',
                        'Eff Util', 'Planned Maint', 'Unplanned', 'Internal Delays', 'External Delays',
                        'Lump Tonnes', 'Fines Tonnes', 'Hours/Day']

    for col, header in enumerate(forecast_headers, 1):
        cell = ws_forecast.cell(row=4, column=col, value=header)
        cell.font = subheader_font
        cell.fill = forecast_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    data_start_row = 5
    for row_idx, (_, row_data) in enumerate(df_forecast.iterrows(), data_start_row):
        ws_forecast.cell(row=row_idx, column=1, value=row_data['Month']).border = thin_border
        ws_forecast.cell(row=row_idx, column=2, value=row_data['Tonnes']).border = thin_border
        ws_forecast.cell(row=row_idx, column=3, value=row_data['Run Hours']).border = thin_border
        ws_forecast.cell(row=row_idx, column=4, value=row_data['TPH']).border = thin_border
        ws_forecast.cell(row=row_idx, column=5, value=row_data['Availability']).border = thin_border
        ws_forecast.cell(row=row_idx, column=5).number_format = '0.00%'
        ws_forecast.cell(row=row_idx, column=6, value=row_data['Utilisation']).border = thin_border
        ws_forecast.cell(row=row_idx, column=6).number_format = '0.00%'
        ws_forecast.cell(row=row_idx, column=7, value=row_data['Effective Utilisation']).border = thin_border
        ws_forecast.cell(row=row_idx, column=7).number_format = '0.00%'

        if 'Planned Maint' in row_data:
            ws_forecast.cell(row=row_idx, column=8, value=row_data['Planned Maint']).border = thin_border
        if 'Unplanned' in row_data:
            ws_forecast.cell(row=row_idx, column=9, value=row_data['Unplanned']).border = thin_border
        if 'Internal Delays' in row_data:
            ws_forecast.cell(row=row_idx, column=10, value=row_data['Internal Delays']).border = thin_border
        if 'External Delays' in row_data:
            ws_forecast.cell(row=row_idx, column=11, value=row_data['External Delays']).border = thin_border

        # Lump = Tonnes * lump_split (formula)
        ws_forecast.cell(row=row_idx, column=12, value=f"=B{row_idx}*{lump_split}").border = thin_border
        # Fines = Tonnes * fines_split (formula)
        ws_forecast.cell(row=row_idx, column=13, value=f"=B{row_idx}*{fines_split}").border = thin_border
        # Hours/Day = Run Hours / days in month (approx 30)
        ws_forecast.cell(row=row_idx, column=14, value=f"=C{row_idx}/30").border = thin_border

    data_end_row = data_start_row + num_forecast_months - 1

    # TOTALS/AVERAGES row with formulas
    totals_row = data_end_row + 2
    ws_forecast.cell(row=totals_row, column=1, value="TOTAL/AVG").font = subheader_font
    ws_forecast.cell(row=totals_row, column=1).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=1).border = thin_border

    # Tonnes - SUM
    ws_forecast.cell(row=totals_row, column=2, value=f"=SUM(B{data_start_row}:B{data_end_row})")
    ws_forecast.cell(row=totals_row, column=2).font = formula_font
    ws_forecast.cell(row=totals_row, column=2).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=2).border = thin_border

    # Run Hours - SUM
    ws_forecast.cell(row=totals_row, column=3, value=f"=SUM(C{data_start_row}:C{data_end_row})")
    ws_forecast.cell(row=totals_row, column=3).font = formula_font
    ws_forecast.cell(row=totals_row, column=3).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=3).border = thin_border

    # TPH - AVERAGE
    ws_forecast.cell(row=totals_row, column=4, value=f"=AVERAGE(D{data_start_row}:D{data_end_row})")
    ws_forecast.cell(row=totals_row, column=4).font = formula_font
    ws_forecast.cell(row=totals_row, column=4).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=4).border = thin_border

    # Availability - AVERAGE
    ws_forecast.cell(row=totals_row, column=5, value=f"=AVERAGE(E{data_start_row}:E{data_end_row})")
    ws_forecast.cell(row=totals_row, column=5).font = formula_font
    ws_forecast.cell(row=totals_row, column=5).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=5).border = thin_border
    ws_forecast.cell(row=totals_row, column=5).number_format = '0.00%'

    # Utilisation - AVERAGE
    ws_forecast.cell(row=totals_row, column=6, value=f"=AVERAGE(F{data_start_row}:F{data_end_row})")
    ws_forecast.cell(row=totals_row, column=6).font = formula_font
    ws_forecast.cell(row=totals_row, column=6).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=6).border = thin_border
    ws_forecast.cell(row=totals_row, column=6).number_format = '0.00%'

    # Eff Util - AVERAGE
    ws_forecast.cell(row=totals_row, column=7, value=f"=AVERAGE(G{data_start_row}:G{data_end_row})")
    ws_forecast.cell(row=totals_row, column=7).font = formula_font
    ws_forecast.cell(row=totals_row, column=7).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=7).border = thin_border
    ws_forecast.cell(row=totals_row, column=7).number_format = '0.00%'

    # Delays - SUM
    for col in range(8, 12):
        ws_forecast.cell(row=totals_row, column=col, value=f"=SUM({col_letter(col)}{data_start_row}:{col_letter(col)}{data_end_row})")
        ws_forecast.cell(row=totals_row, column=col).font = formula_font
        ws_forecast.cell(row=totals_row, column=col).fill = calc_fill
        ws_forecast.cell(row=totals_row, column=col).border = thin_border

    # Lump/Fines - SUM
    ws_forecast.cell(row=totals_row, column=12, value=f"=SUM(L{data_start_row}:L{data_end_row})")
    ws_forecast.cell(row=totals_row, column=12).font = formula_font
    ws_forecast.cell(row=totals_row, column=12).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=12).border = thin_border

    ws_forecast.cell(row=totals_row, column=13, value=f"=SUM(M{data_start_row}:M{data_end_row})")
    ws_forecast.cell(row=totals_row, column=13).font = formula_font
    ws_forecast.cell(row=totals_row, column=13).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=13).border = thin_border

    # Hours/Day - AVERAGE
    ws_forecast.cell(row=totals_row, column=14, value=f"=AVERAGE(N{data_start_row}:N{data_end_row})")
    ws_forecast.cell(row=totals_row, column=14).font = formula_font
    ws_forecast.cell(row=totals_row, column=14).fill = calc_fill
    ws_forecast.cell(row=totals_row, column=14).border = thin_border

    # Formula explanations
    formula_row = totals_row + 2
    ws_forecast.cell(row=formula_row, column=1, value="FORMULA KEY:").font = subheader_font
    ws_forecast.cell(row=formula_row+1, column=1, value="Lump Tonnes").font = value_font
    ws_forecast.cell(row=formula_row+1, column=2, value=f"= Tonnes * {lump_split} (Lump Split %)").font = formula_font
    ws_forecast.cell(row=formula_row+2, column=1, value="Fines Tonnes").font = value_font
    ws_forecast.cell(row=formula_row+2, column=2, value=f"= Tonnes * {fines_split} (Fines Split %)").font = formula_font
    ws_forecast.cell(row=formula_row+3, column=1, value="Hours/Day").font = value_font
    ws_forecast.cell(row=formula_row+3, column=2, value="= Run Hours / 30 (approx days/month)").font = formula_font
    ws_forecast.cell(row=formula_row+4, column=1, value="Eff Utilisation").font = value_font
    ws_forecast.cell(row=formula_row+4, column=2, value="= Availability * Utilisation").font = formula_font

    # Column widths
    col_widths = [10, 14, 12, 10, 12, 12, 10, 14, 12, 14, 14, 14, 14, 10]
    for i, width in enumerate(col_widths, 1):
        ws_forecast.column_dimensions[col_letter(i)].width = width

    # =========================================================================
    # SHEET 2: HISTORICAL DATA (if available)
    # =========================================================================
    if df_historical is not None:
        ws_historical = wb.create_sheet("Historical Data")

        # Header
        ws_historical.merge_cells('A1:P1')
        ws_historical['A1'] = f"HISTORICAL DATA - {site_name}"
        ws_historical['A1'].font = header_font
        ws_historical['A1'].fill = header_fill
        ws_historical['A1'].alignment = Alignment(horizontal='center')

        ws_historical['A2'] = "Period: Jul-24 to Nov-25 (17 months)"
        ws_historical['A2'].font = value_font

        # Column headers
        hist_headers = ['Month', 'Tonnes', 'Run Hours', 'TPH', 'Availability', 'Utilisation',
                        'Eff Util', 'Planned Maint', 'Unplanned', 'Internal Delays', 'External Delays']

        for col, header in enumerate(hist_headers, 1):
            cell = ws_historical.cell(row=4, column=col, value=header)
            cell.font = subheader_font
            cell.fill = historical_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        hist_start_row = 5
        for row_idx, (_, row_data) in enumerate(df_historical.iterrows(), hist_start_row):
            ws_historical.cell(row=row_idx, column=1, value=row_data['Month']).border = thin_border
            ws_historical.cell(row=row_idx, column=2, value=row_data['Tonnes']).border = thin_border
            ws_historical.cell(row=row_idx, column=3, value=row_data['Run Hours']).border = thin_border
            ws_historical.cell(row=row_idx, column=4, value=row_data['TPH']).border = thin_border
            ws_historical.cell(row=row_idx, column=5, value=row_data['Availability']).border = thin_border
            ws_historical.cell(row=row_idx, column=5).number_format = '0.00%'
            ws_historical.cell(row=row_idx, column=6, value=row_data['Utilisation']).border = thin_border
            ws_historical.cell(row=row_idx, column=6).number_format = '0.00%'

            # Effective Utilisation - formula
            ws_historical.cell(row=row_idx, column=7, value=f"=E{row_idx}*F{row_idx}").border = thin_border
            ws_historical.cell(row=row_idx, column=7).number_format = '0.00%'

            if 'Planned Maint' in row_data:
                ws_historical.cell(row=row_idx, column=8, value=row_data['Planned Maint']).border = thin_border
            if 'Unplanned' in row_data:
                ws_historical.cell(row=row_idx, column=9, value=row_data['Unplanned']).border = thin_border
            if 'Internal Delays' in row_data:
                ws_historical.cell(row=row_idx, column=10, value=row_data['Internal Delays']).border = thin_border
            if 'External Delays' in row_data:
                ws_historical.cell(row=row_idx, column=11, value=row_data['External Delays']).border = thin_border

        hist_end_row = hist_start_row + num_hist_months - 1

        # TOTALS/AVERAGES row with formulas
        hist_totals_row = hist_end_row + 2
        ws_historical.cell(row=hist_totals_row, column=1, value="TOTAL/AVG").font = subheader_font
        ws_historical.cell(row=hist_totals_row, column=1).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=1).border = thin_border

        # Tonnes - SUM
        ws_historical.cell(row=hist_totals_row, column=2, value=f"=SUM(B{hist_start_row}:B{hist_end_row})")
        ws_historical.cell(row=hist_totals_row, column=2).font = formula_font
        ws_historical.cell(row=hist_totals_row, column=2).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=2).border = thin_border

        # Run Hours - SUM
        ws_historical.cell(row=hist_totals_row, column=3, value=f"=SUM(C{hist_start_row}:C{hist_end_row})")
        ws_historical.cell(row=hist_totals_row, column=3).font = formula_font
        ws_historical.cell(row=hist_totals_row, column=3).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=3).border = thin_border

        # TPH - AVERAGE
        ws_historical.cell(row=hist_totals_row, column=4, value=f"=AVERAGE(D{hist_start_row}:D{hist_end_row})")
        ws_historical.cell(row=hist_totals_row, column=4).font = formula_font
        ws_historical.cell(row=hist_totals_row, column=4).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=4).border = thin_border

        # Availability - AVERAGE
        ws_historical.cell(row=hist_totals_row, column=5, value=f"=AVERAGE(E{hist_start_row}:E{hist_end_row})")
        ws_historical.cell(row=hist_totals_row, column=5).font = formula_font
        ws_historical.cell(row=hist_totals_row, column=5).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=5).border = thin_border
        ws_historical.cell(row=hist_totals_row, column=5).number_format = '0.00%'

        # Utilisation - AVERAGE
        ws_historical.cell(row=hist_totals_row, column=6, value=f"=AVERAGE(F{hist_start_row}:F{hist_end_row})")
        ws_historical.cell(row=hist_totals_row, column=6).font = formula_font
        ws_historical.cell(row=hist_totals_row, column=6).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=6).border = thin_border
        ws_historical.cell(row=hist_totals_row, column=6).number_format = '0.00%'

        # Eff Util - formula (Avg Availability * Avg Utilisation)
        ws_historical.cell(row=hist_totals_row, column=7, value=f"=E{hist_totals_row}*F{hist_totals_row}")
        ws_historical.cell(row=hist_totals_row, column=7).font = formula_font
        ws_historical.cell(row=hist_totals_row, column=7).fill = calc_fill
        ws_historical.cell(row=hist_totals_row, column=7).border = thin_border
        ws_historical.cell(row=hist_totals_row, column=7).number_format = '0.00%'

        # Delays - SUM
        for col in range(8, 12):
            ws_historical.cell(row=hist_totals_row, column=col, value=f"=SUM({col_letter(col)}{hist_start_row}:{col_letter(col)}{hist_end_row})")
            ws_historical.cell(row=hist_totals_row, column=col).font = formula_font
            ws_historical.cell(row=hist_totals_row, column=col).fill = calc_fill
            ws_historical.cell(row=hist_totals_row, column=col).border = thin_border

        # Column widths
        for i, width in enumerate(col_widths[:11], 1):
            ws_historical.column_dimensions[col_letter(i)].width = width

    # =========================================================================
    # SHEET 3: SUMMARY WITH CALCULATIONS
    # =========================================================================
    ws_summary = wb.create_sheet("Summary & Analysis")

    # Header
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = "CSI MINING SERVICES - TIME USAGE MODEL"
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = header_fill
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    ws_summary['A3'] = f"Project: {site_name}"
    ws_summary['A3'].font = title_font
    ws_summary['A4'] = f"Generated: {datetime.now().strftime('%d/%m/%Y')}"

    # Model Parameters
    ws_summary['A6'] = "MODEL PARAMETERS"
    ws_summary['A6'].font = section_font

    params = [
        ("Contract Tonnes:", contract_tonnes if contract_tonnes else "='Forecast Data'!B" + str(totals_row), "Target production"),
        ("Target TPH:", target_tph if target_tph else "='Forecast Data'!D" + str(totals_row), "Target throughput rate"),
        ("Lump Split:", lump_split, "Proportion of lump product"),
        ("Fines Split:", fines_split, "Proportion of fines product"),
        ("Forecast Months:", num_forecast_months, "Number of months in forecast"),
    ]

    for i, (label, value, desc) in enumerate(params):
        ws_summary.cell(row=7+i, column=1, value=label).font = subheader_font
        if isinstance(value, str) and value.startswith('='):
            ws_summary.cell(row=7+i, column=2, value=value).font = formula_font
        else:
            ws_summary.cell(row=7+i, column=2, value=value).font = value_font
        ws_summary.cell(row=7+i, column=3, value=desc).font = Font(italic=True, color='666666')

    # FY27 FORECAST SUMMARY
    fc_start = 14
    ws_summary.cell(row=fc_start, column=1, value="FY27 FORECAST SUMMARY").font = section_font

    forecast_calcs = [
        ("Total Tonnes", f"='Forecast Data'!B{totals_row}", "SUM of monthly tonnes"),
        ("Total Run Hours", f"='Forecast Data'!C{totals_row}", "SUM of monthly run hours"),
        ("Avg TPH", f"='Forecast Data'!D{totals_row}", "AVERAGE of monthly TPH"),
        ("Avg Availability", f"='Forecast Data'!E{totals_row}", "AVERAGE of monthly availability"),
        ("Avg Utilisation", f"='Forecast Data'!F{totals_row}", "AVERAGE of monthly utilisation"),
        ("Avg Eff Utilisation", f"='Forecast Data'!G{totals_row}", "AVERAGE of monthly eff util"),
        ("Total Lump Tonnes", f"='Forecast Data'!L{totals_row}", "SUM(Tonnes * Lump Split)"),
        ("Total Fines Tonnes", f"='Forecast Data'!M{totals_row}", "SUM(Tonnes * Fines Split)"),
        ("Total Planned Maint", f"='Forecast Data'!H{totals_row}", "SUM of planned maintenance hours"),
        ("Total Unplanned", f"='Forecast Data'!I{totals_row}", "SUM of unplanned downtime hours"),
        ("Total Internal Delays", f"='Forecast Data'!J{totals_row}", "SUM of internal delay hours"),
        ("Total External Delays", f"='Forecast Data'!K{totals_row}", "SUM of external delay hours"),
    ]

    # Headers
    ws_summary.cell(row=fc_start+1, column=1, value="Metric").font = subheader_font
    ws_summary.cell(row=fc_start+1, column=1).fill = forecast_fill
    ws_summary.cell(row=fc_start+1, column=2, value="Value").font = subheader_font
    ws_summary.cell(row=fc_start+1, column=2).fill = forecast_fill
    ws_summary.cell(row=fc_start+1, column=3, value="Calculation").font = subheader_font
    ws_summary.cell(row=fc_start+1, column=3).fill = forecast_fill

    for i, (metric, formula, calc_desc) in enumerate(forecast_calcs, fc_start+2):
        ws_summary.cell(row=i, column=1, value=metric).font = value_font
        ws_summary.cell(row=i, column=1).border = thin_border
        ws_summary.cell(row=i, column=2, value=formula).font = formula_font
        ws_summary.cell(row=i, column=2).border = thin_border
        if 'Availability' in metric or 'Utilisation' in metric:
            ws_summary.cell(row=i, column=2).number_format = '0.00%'
        ws_summary.cell(row=i, column=3, value=calc_desc).font = Font(italic=True, color='666666')

    # HISTORICAL SUMMARY (if available)
    if df_historical is not None:
        hist_summary_start = fc_start + len(forecast_calcs) + 4
        ws_summary.cell(row=hist_summary_start, column=1, value="HISTORICAL SUMMARY (Jul-24 to Nov-25)").font = section_font

        hist_calcs = [
            ("Total Tonnes", f"='Historical Data'!B{hist_totals_row}", "SUM of monthly tonnes"),
            ("Total Run Hours", f"='Historical Data'!C{hist_totals_row}", "SUM of monthly run hours"),
            ("Avg TPH", f"='Historical Data'!D{hist_totals_row}", "AVERAGE of monthly TPH"),
            ("Avg Availability", f"='Historical Data'!E{hist_totals_row}", "AVERAGE of monthly availability"),
            ("Avg Utilisation", f"='Historical Data'!F{hist_totals_row}", "AVERAGE of monthly utilisation"),
            ("Avg Eff Utilisation", f"='Historical Data'!G{hist_totals_row}", "Avg Availability * Avg Utilisation"),
        ]

        # Headers
        ws_summary.cell(row=hist_summary_start+1, column=1, value="Metric").font = subheader_font
        ws_summary.cell(row=hist_summary_start+1, column=1).fill = historical_fill
        ws_summary.cell(row=hist_summary_start+1, column=2, value="Value").font = subheader_font
        ws_summary.cell(row=hist_summary_start+1, column=2).fill = historical_fill
        ws_summary.cell(row=hist_summary_start+1, column=3, value="Calculation").font = subheader_font
        ws_summary.cell(row=hist_summary_start+1, column=3).fill = historical_fill

        for i, (metric, formula, calc_desc) in enumerate(hist_calcs, hist_summary_start+2):
            ws_summary.cell(row=i, column=1, value=metric).font = value_font
            ws_summary.cell(row=i, column=1).border = thin_border
            ws_summary.cell(row=i, column=2, value=formula).font = formula_font
            ws_summary.cell(row=i, column=2).border = thin_border
            if 'Availability' in metric or 'Utilisation' in metric:
                ws_summary.cell(row=i, column=2).number_format = '0.00%'
            ws_summary.cell(row=i, column=3, value=calc_desc).font = Font(italic=True, color='666666')

        # GAP ANALYSIS
        gap_start = hist_summary_start + len(hist_calcs) + 4
    else:
        gap_start = fc_start + len(forecast_calcs) + 4

    ws_summary.cell(row=gap_start, column=1, value="GAP ANALYSIS").font = section_font

    # Contract vs Forecast comparison
    contract_val = contract_tonnes if contract_tonnes else f"='Forecast Data'!B{totals_row}"

    gap_calcs = [
        ("Contract Target", contract_val if isinstance(contract_val, (int, float)) else contract_val, "User-defined target"),
        ("Forecast Total", f"='Forecast Data'!B{totals_row}", "Total forecast tonnes"),
        ("Gap (Tonnes)", f"={contract_val if isinstance(contract_val, (int, float)) else 'B' + str(gap_start+2)}-'Forecast Data'!B{totals_row}", "Contract - Forecast"),
        ("Gap (%)", f"=({contract_val if isinstance(contract_val, (int, float)) else 'B' + str(gap_start+2)}-'Forecast Data'!B{totals_row})/{contract_val if isinstance(contract_val, (int, float)) else 'B' + str(gap_start+2)}", "(Contract - Forecast) / Contract"),
    ]

    ws_summary.cell(row=gap_start+1, column=1, value="Metric").font = subheader_font
    ws_summary.cell(row=gap_start+1, column=1).fill = calc_fill
    ws_summary.cell(row=gap_start+1, column=2, value="Value").font = subheader_font
    ws_summary.cell(row=gap_start+1, column=2).fill = calc_fill
    ws_summary.cell(row=gap_start+1, column=3, value="Calculation").font = subheader_font
    ws_summary.cell(row=gap_start+1, column=3).fill = calc_fill

    for i, (metric, formula, calc_desc) in enumerate(gap_calcs, gap_start+2):
        ws_summary.cell(row=i, column=1, value=metric).font = value_font
        ws_summary.cell(row=i, column=1).border = thin_border
        if isinstance(formula, (int, float)):
            ws_summary.cell(row=i, column=2, value=formula).font = value_font
        else:
            ws_summary.cell(row=i, column=2, value=formula).font = formula_font
        ws_summary.cell(row=i, column=2).border = thin_border
        if 'Gap (%)' in metric:
            ws_summary.cell(row=i, column=2).number_format = '0.00%'
        ws_summary.cell(row=i, column=3, value=calc_desc).font = Font(italic=True, color='666666')

    # Column widths
    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 40

    # =========================================================================
    # SHEET 4: METHODOLOGY
    # =========================================================================
    ws_method = wb.create_sheet("Methodology")

    ws_method.merge_cells('A1:E1')
    ws_method['A1'] = "CALCULATION METHODOLOGY"
    ws_method['A1'].font = header_font
    ws_method['A1'].fill = header_fill

    ws_method['A3'] = "TIME USAGE MODEL CALCULATIONS"
    ws_method['A3'].font = title_font

    methodology = [
        ("Metric", "Formula", "Description"),
        ("Total Tonnes", "=SUM(Monthly Tonnes)", "Sum of all monthly production"),
        ("Total Run Hours", "=SUM(Monthly Run Hours)", "Sum of all productive hours"),
        ("Average TPH", "=AVERAGE(Monthly TPH)", "Mean throughput rate"),
        ("Average Availability", "=AVERAGE(Monthly Availability)", "Mean asset availability"),
        ("Average Utilisation", "=AVERAGE(Monthly Utilisation)", "Mean asset utilisation"),
        ("Effective Utilisation", "=Availability * Utilisation", "Combined effectiveness metric"),
        ("Lump Tonnes", "=Tonnes * Lump Split %", "Premium product volume"),
        ("Fines Tonnes", "=Tonnes * Fines Split %", "Standard product volume"),
        ("Hours per Day", "=Run Hours / Days in Month", "Daily productive hours"),
        ("Gap (Tonnes)", "=Contract Target - Forecast Total", "Shortfall/surplus vs target"),
        ("Gap (%)", "=(Contract - Forecast) / Contract", "Percentage variance"),
        ("", "", ""),
        ("TIME CLASSIFICATION", "", ""),
        ("Calendar Time", "=Days * 24", "Total hours in period"),
        ("Unavailable Time", "=Planned Maint + Unplanned", "Hours asset cannot run"),
        ("Available Time", "=Calendar - Unavailable", "Hours asset can run"),
        ("Unproductive Time", "=Internal + External Delays", "Hours asset doesn't run"),
        ("Productive Time (Run Hours)", "=Available - Unproductive", "Actual running time"),
        ("", "", ""),
        ("KPI DEFINITIONS", "", ""),
        ("Availability %", "=(Calendar - Unavailable) / Calendar", "Asset readiness"),
        ("Utilisation %", "=Run Hours / Available Hours", "Usage of available time"),
        ("Effective Utilisation %", "=Availability * Utilisation", "Overall equipment effectiveness"),
    ]

    for row_idx, (metric, formula, desc) in enumerate(methodology, 5):
        ws_method.cell(row=row_idx, column=1, value=metric)
        ws_method.cell(row=row_idx, column=2, value=formula)
        ws_method.cell(row=row_idx, column=3, value=desc)

        if row_idx == 5 or metric in ["TIME CLASSIFICATION", "KPI DEFINITIONS"]:
            ws_method.cell(row=row_idx, column=1).font = subheader_font
            ws_method.cell(row=row_idx, column=2).font = subheader_font
            ws_method.cell(row=row_idx, column=3).font = subheader_font
            ws_method.cell(row=row_idx, column=1).fill = light_fill
            ws_method.cell(row=row_idx, column=2).fill = light_fill
            ws_method.cell(row=row_idx, column=3).fill = light_fill
        else:
            ws_method.cell(row=row_idx, column=1).font = value_font
            ws_method.cell(row=row_idx, column=2).font = formula_font
            ws_method.cell(row=row_idx, column=3).font = Font(italic=True, color='666666')

    ws_method.column_dimensions['A'].width = 25
    ws_method.column_dimensions['B'].width = 40
    ws_method.column_dimensions['C'].width = 35

    # Save workbook
    wb.save(output_path)
    print(f"Report saved to: {output_path}")

    return {
        'num_forecast_months': num_forecast_months,
        'num_hist_months': num_hist_months,
        'totals_row': totals_row,
    }


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description="Generate CSI TUM Report with calculations")
    parser.add_argument('--site', '-s', required=True, help='Site name')
    parser.add_argument('--forecast', '-f', required=True, help='Path to forecast CSV')
    parser.add_argument('--historical', '-hist', help='Path to historical CSV')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    parser.add_argument('--contract', type=float, help='Contract tonnes target')
    parser.add_argument('--tph', type=float, help='Target TPH')

    args = parser.parse_args()

    create_csi_tum_report(
        site_name=args.site,
        forecast_csv_path=args.forecast,
        historical_csv_path=args.historical,
        output_path=args.output,
        contract_tonnes=args.contract,
        target_tph=args.tph,
    )
